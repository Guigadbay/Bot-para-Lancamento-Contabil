'''
Passos:
- Ler dados da planilha.
- Inserir cada célula de cada linha eum um campo do sistema. 

Bibliotecas:
Openpyxl: Ler planilhas excel.
Pyautogui: Automação de mouse e teclado.

Terminal:
python -m venv guigadbay: Criri o ambiente virtual chamando guigadbay.
./guigadbay/Scripts/activate: Ativei o ambiente virtual.
pip install openpyxl pyautogui: Instalei as bibliotecas.
'''
import openpyxl #Permite ler planilhas.
import pyautogui #Automação de mouse e teclado.

workbook = openpyxl.load_workbook('vendas_de_produtos.xlsx') #Atribui a variavel workbook a função de abrir o arquivo venda_de_produtos.
vendas_sheet = workbook['vendas'] #Atribui a variavel vendas_sheet a função de ler a página vendas do arquivo venda_de_produtos.

#iter_rows função do Openpyxl que permite ler cada linha da planilha seguindo os parametros dos parentesis.
#min_row diz que é para começar a ler no minimo a partir da linha 2.
#linha[0].value definindo que é para a variavel linha ler cada coluna da planilha separadamente, sendo a primeira a coluna 0.
for linha in vendas_sheet.iter_rows(min_row=2):
    #Barra de digitação Cliente.
    pyautogui.click(1498,260,duration=1.5) #Definir a posição do mouse quando esta encima da barra de digitação de Cliente e o tempo de duração do clique.
    pyautogui.write(linha[0].value) #Difinir que é para escrever o conteuno da primeira coluna da planilha na barra de digitação de Cliente.
    #Barra de digitação Produto.
    pyautogui.click(1471,296,duration=1.5)
    pyautogui.write(linha[1].value)
    #Barra de digitação Quantidade.
    pyautogui.click(1480,329,duration=1.5)
    pyautogui.write(str(linha[2].value)) #Definir o que for digitado na barra de digitação de Quantidade, como uma string, porque o pyautogui não entende números.
    #Barra de digitação Categoria do Produto.
    pyautogui.click(1581,361,duration=1.5)
    pyautogui.write(linha[3].value)
    #Botão Salvar.
    pyautogui.click(1385,401,duration=1.5)
    #Botão OK.
    pyautogui.click(816,571,duration=1.5)
'''
Terminal do Prompt de Comando: 
pip install mouseinfo: Intala a biblioteca mouseinfo.
python: Faz o terminal funcionar em python.
from mouseinfo import mouseInfo: Importa p programa mouseInfo.
mouseInfo(): Abre o programa mouseInfo.

Passos:
- Desliga o 3 Sec Button Delay.
- Coloca o mouse encima da barra de digitar de Cliente, Produto, Quantidade, Categoria do Produto, depois no botão salvar, OK e aperta F6 em cada um deles.
- Os numeros que forem salvos no programa, coloca-os no código acima.
'''












